import argparse
import csv
from datetime import datetime
from pathlib import Path
import bcrypt

from models import User

GRADE_BY_CATEGORY_CODE = {
    "8": "Assistant",
    "8c": "Assistant",
    "9a": "Senior",
    "9b": "Assistant manager",
    "10b": "Manager",
    "10c": "Senior manager",
    "11": "Associé",
}


def normalize_email(value):
    if value is None:
        return None
    email = str(value).strip().lower()
    return email or None


def build_full_name(last_name, first_name):
    parts = []
    if first_name:
        parts.append(str(first_name).strip())
    if last_name:
        parts.append(str(last_name).strip())
    return " ".join([p for p in parts if p]) or None


def normalize_category_code(value):
    if value is None:
        return None
    code = str(value).strip()
    return code or None


def resolve_grade_from_category(code):
    if not code:
        return None
    normalized = code.strip().lower()
    return GRADE_BY_CATEGORY_CODE.get(normalized)


def load_tabular_rows(file_path):
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        rows = None
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                with path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.reader(handle, delimiter=";")
                    rows = list(reader)
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            raise ValueError("Impossible de lire le CSV avec les encodages pris en charge.")
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
    else:
        raise ValueError("Format de fichier non pris en charge. Utilisez un CSV ou un fichier Excel.")

    if not rows:
        raise ValueError("Le fichier fourni est vide.")

    headers = rows[0]
    data_rows = rows[1:]
    return headers, data_rows


def import_users_from_file(file_path, default_password, reset_password=False, deactivate_missing=False):
    headers, rows = load_tabular_rows(file_path)
    header_map = {str(h).strip().lower(): idx for idx, h in enumerate(headers) if h}

    required_headers = ["nom", "prenoms", "mail"]
    missing = [h for h in required_headers if h not in header_map]
    if missing:
        raise ValueError(f"Colonnes manquantes dans le fichier: {', '.join(missing)}")

    nom_idx = header_map["nom"]
    prenoms_idx = header_map["prenoms"]
    mail_idx = header_map["mail"]
    grade_idx = header_map.get("grade", header_map.get("code categorie"))
    dep_idx = header_map.get("departement")

    now = datetime.utcnow()

    inserted = 0
    updated = 0
    skipped = 0
    deactivated = 0
    imported_emails = set()

    for row in rows:
        email = normalize_email(row[mail_idx] if mail_idx < len(row) else None)
        if not email:
            skipped += 1
            continue
        imported_emails.add(email)

        last_name = row[nom_idx] if nom_idx < len(row) else None
        first_name = row[prenoms_idx] if prenoms_idx < len(row) else None
        name = build_full_name(last_name, first_name)
        category_code = normalize_category_code(row[grade_idx]) if grade_idx is not None and grade_idx < len(row) else None

        profile_updates = {
            "name": name,
            "last_name": str(last_name).strip() if last_name else None,
            "first_name": str(first_name).strip() if first_name else None,
            "code_categorie": category_code,
            "grade": resolve_grade_from_category(category_code),
            "department": str(row[dep_idx]).strip() if dep_idx is not None and row[dep_idx] else None,
            "is_active": True,
            "updated_at": now,
        }

        existing_user = User.find_by_email(email)
        if existing_user:
            update_doc = {"$set": profile_updates}
            if reset_password:
                update_doc["$set"]["password"] = bcrypt.hashpw(default_password.encode("utf-8"), bcrypt.gensalt())
            User.collection.update_one({"_id": existing_user["_id"]}, update_doc)
            updated += 1
        else:
            user_data = {
                "email": email,
                "password": bcrypt.hashpw(default_password.encode("utf-8"), bcrypt.gensalt()),
                "created_at": now,
                **profile_updates,
            }
            User.collection.insert_one(user_data)
            inserted += 1

    if deactivate_missing:
        result = User.collection.update_many(
            {"email": {"$nin": sorted(imported_emails)}},
            {"$set": {"is_active": False, "updated_at": now}},
        )
        deactivated = result.modified_count

    return inserted, updated, skipped, deactivated


def main():
    parser = argparse.ArgumentParser(
        description="Importer le registre du personnel dans la collection users de MongoDB."
    )
    parser.add_argument("--file", required=True, help="Chemin du fichier CSV ou Excel")
    parser.add_argument(
        "--default-password",
        required=True,
        help="Mot de passe par defaut assigne aux nouveaux utilisateurs",
    )
    parser.add_argument(
        "--reset-password",
        action="store_true",
        help="Reinitialise aussi le mot de passe des utilisateurs deja existants",
    )
    parser.add_argument(
        "--deactivate-missing",
        action="store_true",
        help="Desactive les utilisateurs absents du fichier importe",
    )
    args = parser.parse_args()

    inserted, updated, skipped, deactivated = import_users_from_file(
        file_path=args.file,
        default_password=args.default_password,
        reset_password=args.reset_password,
        deactivate_missing=args.deactivate_missing,
    )
    
    print("Import termine.")
    print(f"- Nouveaux utilisateurs : {inserted}")
    print(f"- Utilisateurs mis a jour : {updated}")
    print(f"- Lignes ignorees (email vide) : {skipped}")
    print(f"- Utilisateurs desactives : {deactivated}")


if __name__ == "__main__":
    main()
